"""
OSA Observatory — collectors/fetcher_usgs_xls.py
Sprint 7 — Avril 2026
Fetcher USGS Mineral Yearbook Africa (XLS/XLSX)
Source : USGS Minerals Yearbook, Volume III, Africa
URL    : https://www.usgs.gov/centers/national-minerals-information-center/africa
Fichiers : myb3-sum-YYYY-africa.xls / myb3-YYYY-africa.xlsx
Indicateurs produits :
  MIN_BAU  — Bauxite (thousand metric tons)
  MIN_CHR  — Chromite (thousand metric tons)
  MIN_COB  — Cobalt (metric tons)
  MIN_COP  — Copper (thousand metric tons)
  MIN_GOL  — Gold (kilograms)
  MIN_IRO  — Iron ore (thousand metric tons)
  MIN_MAN  — Manganese (thousand metric tons)
Usage :
  python collectors/fetcher_usgs_xls.py --dir data/raw/pmin --dry-run
  python collectors/fetcher_usgs_xls.py --dir data/raw/pmin
  python collectors/fetcher_usgs_xls.py --file data/raw/pmin/myb3-sum-2010-africa.xls --dry-run
"""
from __future__ import annotations
import argparse, logging, os, re, sys
from pathlib import Path
import pandas as pd
import psycopg2
from psycopg2.extras import execute_batch
from dotenv import load_dotenv
load_dotenv()

logging.basicConfig(level=os.getenv("OSA_LOG_LEVEL","INFO"),
    format="%(asctime)s | %(levelname)-8s | %(message)s")
log = logging.getLogger("fetcher_usgs_xls")

YEAR_MIN, YEAR_MAX, LAYER_RAW, BATCH_SIZE = 2010, 2024, 1, 500

# ── Mapping pays USGS → ISO3 ──────────────────────────────
COUNTRY_MAP = {
    "Algeria":"DZA","Angola":"AGO","Benin":"BEN","Botswana":"BWA",
    "Burkina Faso":"BFA","Burundi":"BDI","Cabo Verde":"CPV",
    "Cape Verde":"CPV","Cameroon":"CMR","Central African Republic":"CAF",
    "Chad":"TCD","Comoros":"COM","Congo (Brazzaville)":"COG",
    "Congo (Kinshasa)":"COD","Democratic Republic of the Congo":"COD",
    "Republic of the Congo":"COG","Cote d'Ivoire":"CIV",
    "Cote dIvoire":"CIV","Djibouti":"DJI","Egypt":"EGY",
    "Equatorial Guinea":"GNQ","Eritrea":"ERI","Eswatini":"SWZ",
    "Swaziland":"SWZ","Ethiopia":"ETH","Gabon":"GAB","Gambia":"GMB",
    "The Gambia":"GMB","Ghana":"GHA","Guinea":"GIN",
    "Guinea-Bissau":"GNB","Kenya":"KEN","Lesotho":"LSO",
    "Liberia":"LBR","Libya":"LBY","Madagascar":"MDG","Malawi":"MWI",
    "Mali":"MLI","Mauritania":"MRT","Mauritius":"MUS","Morocco":"MAR",
    "Mozambique":"MOZ","Namibia":"NAM","Niger":"NER","Nigeria":"NGA",
    "Rwanda":"RWA","Sao Tome and Principe":"STP","Senegal":"SEN",
    "Seychelles":"SYC","Sierra Leone":"SLE","Somalia":"SOM",
    "South Africa":"ZAF","South Sudan":"SSD","Sudan":"SDN",
    "Tanzania":"TZA","Togo":"TGO","Tunisia":"TUN","Uganda":"UGA",
    "Zambia":"ZMB","Zimbabwe":"ZWE",
}

# ── Mapping colonnes USGS → codes ISA ────────────────────
# Chaque entrée : (keyword_in_header, osa_code, unit_multiplier)
# unit_multiplier : pour convertir en unité standard
COLUMN_MAP = [
    (["bauxite"],                  "MIN_VAL", 1.0),     # thousand metric tons
    (["chromite","chrome"],        "MIN_EXP", 1.0),     # thousand metric tons
    (["cobalt","co content"],      "MIN_RES", 0.001),   # metric tons → thousand
    (["copper","cu content"],      "MIN_EXP", 1.0),     # thousand metric tons
    (["gold","au content"],        "MIN_VAL", 1.0),     # kilograms
    (["iron ore","usable ore"],    "MIN_RES", 1.0),     # thousand metric tons
    (["manganese","mn content"],   "MIN_VAL", 1.0),     # thousand metric tons
]

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("OSA_DB_HOST","localhost"),
        port=int(os.getenv("OSA_DB_PORT",5432)),
        dbname=os.getenv("OSA_DB_NAME","osa_db"),
        user=os.getenv("OSA_DB_USER","postgres"),
        password=os.getenv("OSA_DB_PASS",""),
    )

def extract_year(filepath: str) -> int | None:
    """Extrait l'année depuis le nom de fichier."""
    m = re.search(r'(20\d{2})', Path(filepath).name)
    if m:
        return int(m.group(1))
    return None

def read_table4(filepath: str) -> tuple[list, list]:
    """
    Lit Table 4 (ou T3) depuis un fichier XLS/XLSX.
    Retourne (header_rows, data_rows).
    """
    fp = str(filepath)
    ext = Path(fp).suffix.lower()

    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(fp)
        # Chercher la feuille avec production minérale
        sheet = None
        for name in ["Table 4","T4","T3","Table3"]:
            if name in wb.sheet_names():
                sheet = wb.sheet_by_name(name)
                break
        if sheet is None:
            # Prendre la dernière feuille (souvent Table 4)
            sheet = wb.sheet_by_index(-1)
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        sheet = None
        for name in ["Table 4","T4","T3","Table3"]:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb[wb.sheetnames[-1]]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]

    return rows

def find_header_and_data(rows: list) -> tuple[int, int, list]:
    """
    Trouve la ligne 'Country' (début des données) et les lignes header.
    Retourne (header_start, data_start, rows).
    """
    country_row = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().lower().startswith("country"):
            country_row = i
            break
    if country_row is None:
        raise ValueError("Ligne 'Country' non trouvée")
    return max(0, country_row - 4), country_row + 1, rows

def build_column_mapping(rows: list, header_start: int, data_start: int) -> dict:
    """
    Construit le mapping col_index → (osa_code, multiplier)
    en analysant les lignes header.
    """
    # Concaténer les lignes header en une seule chaîne par colonne
    col_labels = {}
    for i in range(rows[0].__len__()):
        parts = []
        for r in range(header_start, data_start):
            val = rows[r][i] if i < len(rows[r]) else None
            if val and str(val).strip():
                parts.append(str(val).strip().lower())
        col_labels[i] = " ".join(parts)

    # Mapper sur les codes ISA
    mapping = {}
    for col_idx, label in col_labels.items():
        for keywords, osa_code, mult in COLUMN_MAP:
            if any(kw in label for kw in keywords):
                # Ne pas écraser si déjà mappé (garder le premier match)
                if osa_code not in [v[0] for v in mapping.values()]:
                    mapping[col_idx] = (osa_code, mult)
                break
    return mapping

def parse_value(val) -> float | None:
    """Nettoie et convertit une valeur USGS."""
    if val is None:
        return None
    s = str(val).strip()
    # Valeurs manquantes
    if s in ["--", "---", "NA", "W", "e", "", "None", "nan"]:
        return None
    # Supprimer les suffixes de note (lettres)
    s = re.sub(r'[a-zA-Z,\s]+$', '', s).strip()
    s = s.replace(",", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

def parse_file(filepath: str) -> list[dict]:
    """Parse un fichier USGS et retourne une liste de records."""
    year = extract_year(filepath)
    if year is None or not (YEAR_MIN <= year <= YEAR_MAX):
        log.warning("  Année %s hors plage [%d,%d] — ignoré", year, YEAR_MIN, YEAR_MAX)
        return []

    log.info("  Parsing %s (année %d)", Path(filepath).name, year)

    rows = read_table4(filepath)
    header_start, data_start, rows = find_header_and_data(rows)
    col_map = build_column_mapping(rows, header_start, data_start)

    if not col_map:
        log.warning("  Aucune colonne minérale trouvée dans %s", filepath)
        return []

    log.debug("  Colonnes mappées : %s",
              {v[0]: k for k, v in col_map.items()})

    records = []
    for row in rows[data_start:]:
        if not row or not row[0]:
            continue
        country_name = str(row[0]).strip()
        # Nettoyer les suffixes de note (ex: "Angolae" → "Angola")
        country_clean = re.sub(r'[a-z]+$', '', country_name).strip()
        # Essayer les deux formes
        iso3 = COUNTRY_MAP.get(country_name) or COUNTRY_MAP.get(country_clean)
        if not iso3:
            continue

        for col_idx, (osa_code, mult) in col_map.items():
            val_raw = row[col_idx] if col_idx < len(row) else None
            val = parse_value(val_raw)
            if val is None:
                continue
            val_final = val * mult
            if val_final < 0:
                continue
            records.append({
                "indicator_code": osa_code,
                "country_iso3":   iso3,
                "year":           year,
                "raw_value":      val_final,
            })

    log.info("  → %d enregistrements (%d pays)",
             len(records),
             len({r["country_iso3"] for r in records}))
    return records

def insert_to_db(conn, records: list[dict], dry_run: bool = False) -> int:
    """Insère les records en ma.indicator_values."""
    if not records:
        return 0

    df = pd.DataFrame(records)
    valid = pd.read_sql("SELECT iso3 FROM rf.countries", conn)["iso3"].tolist()
    df = df[df["country_iso3"].isin(valid)].copy()

    # Vérifier codes indicateurs
    valid_inds = pd.read_sql(
        "SELECT code FROM rf.indicators WHERE pillar_code='PMIN'", conn
    )["code"].tolist()
    df = df[df["indicator_code"].isin(valid_inds)].copy()

    if df.empty:
        log.warning("  Aucun enregistrement valide après filtrage")
        return 0

    if dry_run:
        for ind, grp in df.groupby("indicator_code"):
            log.info("  [DRY-RUN] %s : %d valeurs | %d pays | années %s-%s",
                     ind, len(grp), grp["country_iso3"].nunique(),
                     grp["year"].min(), grp["year"].max())
        log.info("  [DRY-RUN] Total : %d valeurs non insérées", len(df))
        return len(df)

    sql = """
        INSERT INTO ma.indicator_values
            (indicator_code, country_iso3, year, layer_id,
             raw_value, quality_flag, confidence_score, value_status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """
    batch = [(r["indicator_code"], r["country_iso3"], int(r["year"]),
              LAYER_RAW, float(r["raw_value"]), "OK", 0.9, "OBSERVED")
             for _, r in df.iterrows()]

    with conn.cursor() as cur:
        execute_batch(cur, sql, batch, page_size=BATCH_SIZE)
    conn.commit()
    log.info("  → %d insérés", len(batch))
    return len(batch)

def main():
    parser = argparse.ArgumentParser(description="OSA Fetcher USGS Mineral Yearbook")
    parser.add_argument("--dir",     type=str, default=None,
                        help="Dossier contenant les fichiers USGS")
    parser.add_argument("--file",    type=str, default=None,
                        help="Fichier USGS unique à traiter")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.dir and not args.file:
        parser.print_help()
        sys.exit(1)

    log.info("="*60)
    log.info("OSA Fetcher USGS Mineral Yearbook")
    log.info("  Dry-run : %s", args.dry_run)
    log.info("="*60)

    # Collecter les fichiers à traiter
    files = []
    if args.file:
        files = [args.file]
    else:
        d = Path(args.dir)
        files = sorted([
            str(f) for f in d.iterdir()
            if f.suffix.lower() in [".xls", ".xlsx"]
            and "africa" in f.name.lower()
            and "areacodes" not in f.name.lower()
            and "elements" not in f.name.lower()
            and "itemcodes" not in f.name.lower()
        ])

    log.info("  Fichiers trouvés : %d", len(files))
    if not files:
        log.warning("  Aucun fichier USGS trouvé.")
        sys.exit(1)

    # Parser tous les fichiers
    all_records = []
    for fp in files:
        try:
            records = parse_file(fp)
            all_records.extend(records)
        except Exception as e:
            log.error("  Erreur sur %s : %s", fp, e)

    log.info("Total enregistrements préparés : %d", len(all_records))

    if not all_records:
        log.warning("Aucune donnée extraite.")
        sys.exit(1)

    # Déduplication (garder première valeur en cas de doublon)
    seen = set()
    unique_records = []
    for r in all_records:
        key = (r["indicator_code"], r["country_iso3"], r["year"])
        if key not in seen:
            seen.add(key)
            unique_records.append(r)
    log.info("Après déduplication : %d enregistrements", len(unique_records))

    # Insertion
    conn = get_pg_conn()
    try:
        n = insert_to_db(conn, unique_records, dry_run=args.dry_run)
    finally:
        conn.close()

    # Rapport
    df = pd.DataFrame(unique_records)
    print("\n" + "="*60)
    print("RAPPORT USGS MINERAL YEARBOOK")
    print("="*60)
    print(f"Mode     : {'DRY-RUN' if args.dry_run else 'COLLECT'}")
    print(f"Fichiers : {len(files)}")
    print(f"Préparés : {len(unique_records)}")
    print(f"Insérés  : {n}")
    if not df.empty:
        print("\nPar indicateur :")
        for ind, grp in df.groupby("indicator_code"):
            print(f"  {ind:<12} : {len(grp):>5} val | "
                  f"{grp['country_iso3'].nunique()} pays | "
                  f"{grp['year'].min()}-{grp['year'].max()}")
    print("="*60)

if __name__ == "__main__":
    main()
