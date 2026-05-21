"""
============================================================
OSA Observatory — audit_pipeline.py
Sprint 8 — Mai 2026

Audit complet des vices cachés dans le pipeline L1/L2/L3.
Détecte : doublons, manquants, hors bornes, confidence NULL,
          indicateurs non propagés, source_id manquants.

Usage :
  python audit_pipeline.py
  python audit_pipeline.py --year 2024
  python audit_pipeline.py --detail
============================================================
"""
from __future__ import annotations
import argparse
import os
import sys
from datetime import datetime
from pathlib import Path
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

load_dotenv()

# ── Connexion ─────────────────────────────────────────────
def get_conn():
    return psycopg2.connect(
        host     = os.getenv("OSA_DB_HOST", "127.0.0.1"),
        port     = int(os.getenv("OSA_DB_PORT", 5432)),
        dbname   = os.getenv("OSA_DB_NAME", "osa_db"),
        user     = os.getenv("OSA_DB_USER", "postgres"),
        password = os.getenv("OSA_DB_PASS", os.getenv("OSA_DB_PASSWORD", "")),
    )

# ── Couleurs terminal ─────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
BLUE   = "\033[94m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def ok(msg):    print(f"  {GREEN}✓  {msg}{RESET}")
def alerte(msg): print(f"  {RED}✗  {msg}{RESET}")
def warn(msg):  print(f"  {YELLOW}⚠  {msg}{RESET}")
def info(msg):  print(f"  {BLUE}   {msg}{RESET}")

# ── Checks ────────────────────────────────────────────────

CHECKS = [

    {
        "code": "DOUBLONS_L1",
        "label": "Doublons en L1 (même indicateur/pays/année/layer)",
        "query": """
            SELECT COUNT(*) AS n
            FROM ma.indicator_values
            WHERE layer_id = 1
              AND id NOT IN (
                  SELECT MIN(id) FROM ma.indicator_values
                  WHERE layer_id = 1
                  GROUP BY indicator_code, country_iso3, year, layer_id
              )
        """,
        "detail_query": """
            SELECT indicator_code,
                   COUNT(*) - COUNT(DISTINCT country_iso3||year::text) AS doublons
            FROM ma.indicator_values
            WHERE layer_id = 1
            GROUP BY indicator_code
            HAVING COUNT(*) > COUNT(DISTINCT country_iso3||year::text)
            ORDER BY doublons DESC
            LIMIT 10
        """,
        "seuil": 0,
        "niveau": "CRITIQUE",
    },

    {
        "code": "L1_SOURCE_NULL",
        "label": "Lignes L1 sans source_id (fetcher WB non corrigé)",
        "query": """
            SELECT COUNT(*) AS n
            FROM ma.indicator_values
            WHERE layer_id = 1 AND source_id IS NULL
        """,
        "detail_query": """
            SELECT indicator_code, COUNT(*) AS nb
            FROM ma.indicator_values
            WHERE layer_id = 1 AND source_id IS NULL
            GROUP BY indicator_code
            ORDER BY nb DESC
            LIMIT 10
        """,
        "seuil": 0,
        "niveau": "ATTENTION",
    },

    {
        "code": "L1_SANS_L2",
        "label": "Indicateurs présents en L1 mais absents en L2",
        "query": """
            SELECT COUNT(DISTINCT indicator_code) AS n
            FROM (
                SELECT DISTINCT indicator_code FROM ma.indicator_values WHERE layer_id = 1
                EXCEPT
                SELECT DISTINCT indicator_code FROM ma.indicator_values WHERE layer_id = 2
            ) x
        """,
        "detail_query": """
            SELECT DISTINCT indicator_code
            FROM ma.indicator_values WHERE layer_id = 1
            EXCEPT
            SELECT DISTINCT indicator_code
            FROM ma.indicator_values WHERE layer_id = 2
            ORDER BY indicator_code
        """,
        "seuil": 0,
        "niveau": "CRITIQUE",
    },

    {
        "code": "L2_PAYS_INSUFFISANT",
        "label": "Indicateurs avec moins de pays en L2 qu'en L1",
        "query": """
            SELECT COUNT(*) AS n
            FROM (
                SELECT indicator_code,
                    COUNT(DISTINCT CASE WHEN layer_id=1 THEN country_iso3 END) AS pays_l1,
                    COUNT(DISTINCT CASE WHEN layer_id=2 THEN country_iso3 END) AS pays_l2
                FROM ma.indicator_values
                WHERE year BETWEEN 2010 AND 2024
                GROUP BY indicator_code
                HAVING COUNT(DISTINCT CASE WHEN layer_id=2 THEN country_iso3 END) <
                       COUNT(DISTINCT CASE WHEN layer_id=1 THEN country_iso3 END)
            ) x
        """,
        "detail_query": """
            SELECT indicator_code,
                COUNT(DISTINCT CASE WHEN layer_id=1 THEN country_iso3 END) AS pays_l1,
                COUNT(DISTINCT CASE WHEN layer_id=2 THEN country_iso3 END) AS pays_l2,
                COUNT(DISTINCT CASE WHEN layer_id=1 THEN country_iso3 END) -
                COUNT(DISTINCT CASE WHEN layer_id=2 THEN country_iso3 END) AS manquants
            FROM ma.indicator_values
            WHERE year BETWEEN 2010 AND 2024
            GROUP BY indicator_code
            HAVING COUNT(DISTINCT CASE WHEN layer_id=2 THEN country_iso3 END) <
                   COUNT(DISTINCT CASE WHEN layer_id=1 THEN country_iso3 END)
            ORDER BY manquants DESC
            LIMIT 15
        """,
        "seuil": 0,
        "niveau": "CRITIQUE",
    },

    {
        "code": "L2_SANS_L3",
        "label": "Indicateurs présents en L2 mais absents en L3",
        "query": """
            SELECT COUNT(DISTINCT indicator_code) AS n
            FROM (
                SELECT DISTINCT indicator_code FROM ma.indicator_values WHERE layer_id = 2
                EXCEPT
                SELECT DISTINCT indicator_code FROM ma.indicator_values WHERE layer_id = 3
            ) x
            WHERE indicator_code NOT IN (
                SELECT DISTINCT indicator_code FROM rf.normalization_bounds
                WHERE min_value = max_value OR min_value IS NULL
            )
        """,
        "detail_query": """
            SELECT DISTINCT iv.indicator_code,
                CASE
                    WHEN nb.indicator_code IS NOT NULL
                    THEN 'Exclu du gel (min=max) — indicateur binaire ou sans variance. Normal.'
                    ELSE 'Absent de L3 sans justification connue — vérifier normalize_indicator()'
                END AS justification
            FROM ma.indicator_values iv
            LEFT JOIN (
                SELECT DISTINCT indicator_code FROM rf.normalization_bounds
                WHERE min_value = max_value OR min_value IS NULL
            ) nb ON nb.indicator_code = iv.indicator_code
            WHERE iv.layer_id = 2
              AND iv.indicator_code NOT IN (
                  SELECT indicator_code FROM ma.indicator_values WHERE layer_id = 3
              )
            ORDER BY justification, iv.indicator_code
        """,
        "seuil": 0,
        "niveau": "ATTENTION",
    },

    {
        "code": "L3_CONFIDENCE_NULL",
        "label": "Lignes L3 avec confidence_score NULL",
        "query": """
            SELECT COUNT(*) AS n
            FROM ma.indicator_values
            WHERE layer_id = 3 AND confidence_score IS NULL
        """,
        "detail_query": """
            SELECT indicator_code, COUNT(*) AS nb_null
            FROM ma.indicator_values
            WHERE layer_id = 3 AND confidence_score IS NULL
            GROUP BY indicator_code
            ORDER BY nb_null DESC
            LIMIT 10
        """,
        "seuil": 0,
        "niveau": "CRITIQUE",
    },

    {
        "code": "L3_HORS_BORNES",
        "label": "Lignes L3 avec processed_value hors [0, 1]",
        "query": """
            SELECT COUNT(*) AS n
            FROM ma.indicator_values
            WHERE layer_id = 3
              AND (processed_value < 0 OR processed_value > 1)
        """,
        "detail_query": """
            SELECT indicator_code,
                   ROUND(MIN(processed_value)::numeric, 4) AS min_val,
                   ROUND(MAX(processed_value)::numeric, 4) AS max_val,
                   COUNT(*) AS nb
            FROM ma.indicator_values
            WHERE layer_id = 3
              AND (processed_value < 0 OR processed_value > 1)
            GROUP BY indicator_code
            ORDER BY nb DESC
            LIMIT 10
        """,
        "seuil": 0,
        "niveau": "ATTENTION",
    },

    {
        "code": "INDICATEURS_ACTIFS_SANS_L3",
        "label": "Indicateurs actifs (non COMPUTED) sans aucune donnée L3",
        "query": """
            SELECT COUNT(*) AS n
            FROM rf.indicators i
            WHERE i.is_active = TRUE
              AND i.imputation_regime != 'COMPUTED'
              AND EXISTS (
                  SELECT 1 FROM ma.indicator_values
                  WHERE indicator_code = i.code AND layer_id = 1
              )
              AND NOT EXISTS (
                  SELECT 1 FROM ma.indicator_values
                  WHERE indicator_code = i.code AND layer_id = 3
              )
        """,
        "detail_query": """
            SELECT i.code, i.pillar_code, i.imputation_regime,
                (SELECT COUNT(*) FROM ma.indicator_values
                 WHERE indicator_code = i.code AND layer_id = 1) AS nb_l1,
                CASE
                    WHEN EXISTS (
                        SELECT 1 FROM ma.indicator_values
                        WHERE indicator_code = i.code AND layer_id = 1
                    ) THEN 'ANOMALIE — données L1 présentes mais pas en L3. Vérifier normalize_indicator().'
                    ELSE 'NON COLLECTE — indicateur planifié mais source non encore intégrée. Normal pour sprints futurs.'
                END AS justification
            FROM rf.indicators i
            WHERE i.is_active = TRUE
              AND i.imputation_regime != 'COMPUTED'
              AND NOT EXISTS (
                  SELECT 1 FROM ma.indicator_values
                  WHERE indicator_code = i.code AND layer_id = 3
              )
            ORDER BY justification DESC, i.pillar_code, i.code
        """,
        "seuil": 0,
        "niveau": "ATTENTION",
    },


    {
        "code": "DOCTRINE_COMPLIANCE",
        "label": "Indicateurs actifs non conformes Doctrine OSA v1",
        "query": """
            SELECT COUNT(*) AS n
            FROM rf.indicators
            WHERE is_active = TRUE
              AND doctrine_compliance_flag = FALSE
        """,
        "detail_query": """
            SELECT code, pillar_code, imputation_regime,
                   'Non conforme Doctrine OSA v1 — proxy comportemental requis' AS justification
            FROM rf.indicators
            WHERE is_active = TRUE
              AND doctrine_compliance_flag = FALSE
            ORDER BY pillar_code, code
        """,
        "seuil": 0,
        "niveau": "ATTENTION",
    },

    {
        "code": "AMAR_LOW_CONFIDENCE",
        "label": "Pays LOW_CONFIDENCE dans AMAR 2024",
        "query": """
            SELECT COUNT(*) AS n
            FROM ma.v_p7i_amar_dashboard
            WHERE year = 2024 AND risk_band = 'LOW_CONFIDENCE'
        """,
        "detail_query": """
            SELECT country_iso3,
                   ROUND(risk_score::numeric, 3) AS score,
                   ROUND(confidence_score::numeric, 3) AS conf
            FROM ma.v_p7i_amar_dashboard
            WHERE year = 2024 AND risk_band = 'LOW_CONFIDENCE'
            ORDER BY confidence_score
            LIMIT 15
        """,
        "seuil": 10,
        "niveau": "ATTENTION",
    },

]


# ── Export Excel ──────────────────────────────────────────

def export_excel(results: list, path: str):
    """Génère un rapport Excel — une feuille par check."""
    if not HAS_EXCEL:
        print("  openpyxl non installé — pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # Couleurs
    C_OK      = PatternFill("solid", fgColor="D4EDDA")
    C_CRIT    = PatternFill("solid", fgColor="F8D7DA")
    C_WARN    = PatternFill("solid", fgColor="FFF3CD")
    C_HEADER  = PatternFill("solid", fgColor="1B3A5C")
    F_WHITE   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    F_BOLD    = Font(bold=True, name="Arial", size=10)
    F_NORMAL  = Font(name="Arial", size=10)

    # ── Feuille Synthèse ──────────────────────────────────
    ws = wb.active
    ws.title = "Synthese"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18

    headers = ["Code", "Libellé", "Anomalies", "Niveau", "Statut"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = C_HEADER
        cell.font = F_WHITE
        cell.alignment = Alignment(horizontal="center")

    # Titre
    ws.insert_rows(1)
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"OSA Observatory — Rapport d'audit pipeline — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.fill = C_HEADER
    title_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    title_cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 3):
        statut = "✓ OK" if r["statut"] == "OK" else ("✗ CRITIQUE" if r["niveau"] == "CRITIQUE" else "⚠ ATTENTION")
        fill = C_OK if r["statut"] == "OK" else (C_CRIT if r["niveau"] == "CRITIQUE" else C_WARN)
        vals = [r["code"], r["label"], r["n"], r["niveau"], statut]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = fill
            cell.font = F_BOLD if col == 1 else F_NORMAL
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left")

    # ── Feuilles détail ───────────────────────────────────
    for r in results:
        if not r.get("detail_rows") or r["statut"] == "OK":
            continue
        sheet_name = r["code"][:31]
        ws2 = wb.create_sheet(title=sheet_name)
        ws2.merge_cells("A1:Z1")
        ws2["A1"] = f"{r['code']} — {r['label']} ({r['n']} anomalie(s))"
        ws2["A1"].fill = C_CRIT if r["niveau"] == "CRITIQUE" else C_WARN
        ws2["A1"].font = F_WHITE if r["niveau"] == "CRITIQUE" else Font(bold=True, name="Arial", size=11)

        if r["detail_rows"]:
            cols = list(r["detail_rows"][0].keys())
            for col, c in enumerate(cols, 1):
                cell = ws2.cell(row=2, column=col, value=c)
                cell.fill = C_HEADER
                cell.font = F_WHITE
                ws2.column_dimensions[get_column_letter(col)].width = max(20, len(str(c)) + 4)

            for row_i, row in enumerate(r["detail_rows"], 3):
                for col, c in enumerate(cols, 1):
                    ws2.cell(row=row_i, column=col, value=str(row.get(c, ""))).font = F_NORMAL

    wb.save(path)
    print(f"  Excel exporté : {path}")


# ── Runner ────────────────────────────────────────────────

def run_audit(detail: bool = False, excel_path: str = None):
    print()
    print(f"{BOLD}{'='*62}")
    print(f"  OSA Observatory — Audit pipeline")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*62}{RESET}")
    print()

    try:
        conn = get_conn()
    except Exception as e:
        alerte(f"Connexion DB impossible : {e}")
        sys.exit(1)

    alertes = 0
    attentions = 0
    results = []

    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        for check in CHECKS:
            cur.execute(check["query"])
            row = cur.fetchone()
            n = int(row["n"]) if row else 0
            seuil = check["seuil"]
            niveau = check["niveau"]
            detail_rows = []

            if n <= seuil:
                ok(f"{check['code']:35s} {check['label']}")
                statut = "OK"
            elif niveau == "CRITIQUE":
                alerte(f"{check['code']:35s} {check['label']} → {n}")
                alertes += 1
                statut = "CRITIQUE"
                if (detail or excel_path) and check.get("detail_query"):
                    cur.execute(check["detail_query"])
                    detail_rows = [dict(r) for r in cur.fetchall()]
                    if detail:
                        for r in detail_rows:
                            info(str(r))
            else:
                warn(f"{check['code']:35s} {check['label']} → {n}")
                attentions += 1
                statut = "ATTENTION"
                if (detail or excel_path) and check.get("detail_query"):
                    cur.execute(check["detail_query"])
                    detail_rows = [dict(r) for r in cur.fetchall()]
                    if detail:
                        for r in detail_rows:
                            info(str(r))

            results.append({
                "code": check["code"],
                "label": check["label"],
                "n": n,
                "niveau": niveau,
                "statut": statut,
                "detail_rows": detail_rows,
            })

    conn.close()

    if excel_path:
        export_excel(results, excel_path)

    print()
    print(f"{BOLD}{'='*62}")
    if alertes == 0 and attentions == 0:
        print(f"  {GREEN}AUDIT OK — Aucun problème détecté{RESET}")
    else:
        if alertes > 0:
            print(f"  {RED}{alertes} CRITIQUE(S) — action immédiate requise{RESET}")
        if attentions > 0:
            print(f"  {YELLOW}{attentions} ATTENTION(S) — à surveiller{RESET}")
    print(f"{BOLD}{'='*62}{RESET}")
    print()

    return alertes


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="OSA Pipeline Audit")
    parser.add_argument("--detail", action="store_true",
                        help="Afficher le détail des anomalies dans le terminal")
    parser.add_argument("--excel", type=str, default=None,
                        help="Chemin du fichier Excel de sortie (ex: audit_20260520.xlsx)")
    args = parser.parse_args()

    # Chemin Excel — date automatique si --excel sans valeur
    excel_path = args.excel
    if excel_path is None and "--excel" in sys.argv:
        excel_path = f"logs/audit_pipeline_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif excel_path is not None and not excel_path.endswith('.xlsx'):
        excel_path = f"logs/audit_pipeline_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    n_alertes = run_audit(detail=args.detail, excel_path=excel_path)
    sys.exit(1 if n_alertes > 0 else 0)
