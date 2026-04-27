"""
OSA Observatory
collectors/fetcher_sipri_milex.py -- Ingestion SIPRI Military Expenditure

Indicateur integre :
  MIL_DEP -- Depenses militaires en % du PIB
             Source : SIPRI Milex Database, feuille "Share of GDP"
             Valeurs brutes : ex. 2.5 = 2.5% du PIB
             Normalisation [0,1] : effectuee par normalize_indicator (L3)

Telechargement manuel :
  https://www.sipri.org/databases/milex
  Fichier : SIPRI-Milex-data-1949-2024.xlsx
  Placer dans : data/sipri/SIPRI-Milex-data-1949-2024.xlsx

Usage :
  python collectors/fetcher_sipri_milex.py --file data/sipri/SIPRI-Milex-data-1949-2024.xlsx --dry-run
  python collectors/fetcher_sipri_milex.py --file data/sipri/SIPRI-Milex-data-1949-2024.xlsx
"""

import argparse
import logging
import os
import sys

import openpyxl
import psycopg2
from psycopg2.extras import execute_batch
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=os.getenv("OSA_LOG_LEVEL", "INFO"),
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)
log = logging.getLogger("fetcher_sipri")

# ── Constantes ────────────────────────────────────────────────
SHEET_NAME  = "Share of GDP"      # feuille utilisee
OSA_CODE    = "MIL_DEP"           # indicateur OSA cible
YEAR_FROM   = 2010
YEAR_TO     = 2024
LAYER_RAW   = 1
BATCH_SIZE  = 500
HEADER_ROW  = 5                   # ligne 5 (index 0) contient les annees
DATA_START  = 6                   # donnees a partir de la ligne 6

# Valeurs SIPRI indiquant donnee manquante
MISSING_VALUES = {"...", ". .", "xxx", "", "None", None}

# Sous-regions a ignorer (pas des pays)
SUBREGIONS = {
    "Africa", "North Africa", "Sub-Saharan Africa", "sub-Saharan Africa",
    "Central Africa", "East Africa", "Southern Africa", "West Africa",
    "Europe", "Americas", "Asia & Oceania", "Middle East",
}

# Mapping nom SIPRI -> ISO3
# Seuls les cas ambigus ou non standards sont listés
# Les autres sont résolus via la table rf.countries
SIPRI_NAME_TO_ISO3 = {
    "Egypt":                      "EGY",
    "Algeria":                    "DZA",
    "Libya":                      "LBY",
    "Morocco":                    "MAR",
    "Tunisia":                    "TUN",
    "Angola":                     "AGO",
    "Benin":                      "BEN",
    "Botswana":                   "BWA",
    "Burkina Faso":               "BFA",
    "Burundi":                    "BDI",
    "Cameroon":                   "CMR",
    "Cape Verde":                 "CPV",
    "Central African Republic":   "CAF",
    "Chad":                       "TCD",
    "Congo, DR":                  "COD",
    "Congo, Republic":            "COG",
    "Cote d'Ivoire":              "CIV",
    "Djibouti":                   "DJI",
    "Equatorial Guinea":          "GNQ",
    "Eritrea":                    "ERI",
    "Ethiopia":                   "ETH",
    "Gabon":                      "GAB",
    "Gambia, The":                "GMB",
    "Ghana":                      "GHA",
    "Guinea":                     "GIN",
    "Guinea-Bissau":              "GNB",
    "Kenya":                      "KEN",
    "Lesotho":                    "LSO",
    "Liberia":                    "LBR",
    "Madagascar":                 "MDG",
    "Malawi":                     "MWI",
    "Mali":                       "MLI",
    "Mauritania":                 "MRT",
    "Mauritius":                  "MUS",
    "Mozambique":                 "MOZ",
    "Namibia":                    "NAM",
    "Niger":                      "NER",
    "Nigeria":                    "NGA",
    "Rwanda":                     "RWA",
    "Senegal":                    "SEN",
    "Seychelles":                 "SYC",
    "Sierra Leone":               "SLE",
    "Somalia":                    "SOM",
    "South Africa":               "ZAF",
    "South Sudan":                "SSD",
    "Sudan":                      "SDN",
    "Eswatini":                   "SWZ",
    "Tanzania":                   "TZA",
    "Togo":                       "TGO",
    "Uganda":                     "UGA",
    "Zambia":                     "ZMB",
    "Zimbabwe":                   "ZWE",
}


# ── Connexion PostgreSQL ──────────────────────────────────────
def get_conn():
    return psycopg2.connect(
        host=os.getenv("OSA_DB_HOST", "localhost"),
        port=int(os.getenv("OSA_DB_PORT", 5432)),
        dbname=os.getenv("OSA_DB_NAME", "osa_db"),
        user=os.getenv("OSA_DB_USER", "osa_user"),
        password=os.getenv("OSA_DB_PASS", ""),
    )


def get_method_version(conn):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM ma.indicator_method_versions ORDER BY id DESC LIMIT 1"
        )
        row = cur.fetchone()
        return row[0] if row else 1


def get_african_iso3(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT iso3 FROM rf.countries WHERE iso3 IS NOT NULL")
        return {r[0] for r in cur.fetchall()}


# ── Chargement XLSX ───────────────────────────────────────────
def load_xlsx(filepath):
    """
    Charge la feuille 'Share of GDP' du fichier SIPRI Milex.
    Retourne la liste des lignes brutes.
    """
    if not os.path.exists(filepath):
        log.error("Fichier introuvable : %s", filepath)
        sys.exit(1)

    log.info("Chargement : %s", filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        log.error("Feuille '%s' introuvable. Feuilles disponibles : %s",
                  SHEET_NAME, wb.sheetnames)
        sys.exit(1)

    ws   = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    log.info("Feuille '%s' chargee : %d lignes", SHEET_NAME, len(rows))
    return rows


# ── Parsing et preparation ────────────────────────────────────
def prepare_records(rows, african_iso3, method_version):
    """
    Parse les lignes SIPRI et produit les tuples d'insertion.

    Logique :
    - Ligne 5 (index 0) = header avec les annees
    - Lignes suivantes = pays avec leurs valeurs
    - On reste en mode 'in_africa' entre "Africa" et la prochaine region
    - Valeurs manquantes ("...", None) ignorees
    - Valeurs non numeriques ignorees
    """
    # Trouver les indices des colonnes pour 2010-2024
    header = rows[HEADER_ROW]
    year_indices = {}
    for i, val in enumerate(header):
        try:
            y = int(val)
            if YEAR_FROM <= y <= YEAR_TO:
                year_indices[y] = i
        except (TypeError, ValueError):
            pass

    if not year_indices:
        log.error("Aucune colonne annee trouvee dans la ligne %d", HEADER_ROW)
        sys.exit(1)

    log.info("Colonnes annees trouvees : %d -> %d (%d colonnes)",
             min(year_indices), max(year_indices), len(year_indices))

    records         = []
    in_africa       = False
    pays_couverts   = set()
    pays_sans_iso3  = set()
    n_missing       = 0

    for row in rows[DATA_START:]:
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()

        # Detecter les blocs regionaux
        if name in ("Europe", "Americas", "Asia & Oceania", "Middle East"):
            in_africa = False
            continue
        if name == "Africa":
            in_africa = True
            continue
        # Egypt est classee Middle East dans SIPRI mais fait partie de l'OSA
        if not in_africa and name not in ("Egypt",):
            continue

        # Ignorer les sous-regions
        if name in SUBREGIONS:
            continue

        # Resoudre ISO3
        iso3 = SIPRI_NAME_TO_ISO3.get(name)
        if not iso3:
            pays_sans_iso3.add(name)
            continue

        # Verifier que le pays est dans le referentiel OSA
        if iso3 not in african_iso3:
            log.debug("Pays hors referentiel OSA : %s (%s)", name, iso3)
            continue

        # Parser les valeurs annuelles
        for year, col_idx in year_indices.items():
            if col_idx >= len(row):
                continue

            raw = row[col_idx]

            # Valeur manquante
            if raw in MISSING_VALUES or str(raw).strip() in MISSING_VALUES:
                n_missing += 1
                continue

            try:
                val = float(raw)
            except (TypeError, ValueError):
                n_missing += 1
                continue

            # Toutes les valeurs SIPRI Share of GDP sont en decimales
            # 0.031 = 3.1% du PIB -- conversion obligatoire
            val = val * 100

            # Sanity check -- depenses militaires entre 0 et 25% PIB
            if val < 0 or val > 25:
                log.warning("Valeur hors plage MIL_DEP : %s %d -> %.3f%% PIB -- ignoree",
                            iso3, year, val)
                continue

            records.append((
                OSA_CODE,        # indicator_code
                iso3,            # country_iso3
                year,            # year
                LAYER_RAW,       # layer_id = 1
                val,             # raw_value : % PIB brut
                None,            # processed_value : calcule par normalize_indicator
                method_version,  # method_version_id
                "OK",            # quality_flag
            ))
            pays_couverts.add(iso3)

    # Rapport
    log.info("Enregistrements prepares : %d", len(records))
    log.info("Pays couverts : %d", len(pays_couverts))
    log.info("Valeurs manquantes ignorees : %d", n_missing)

    if pays_sans_iso3:
        log.warning("Pays SIPRI sans mapping ISO3 (%d) : %s",
                    len(pays_sans_iso3), ", ".join(sorted(pays_sans_iso3)))

    manquants = african_iso3 - pays_couverts
    if manquants:
        log.warning("Pays africains OSA absents du fichier SIPRI (%d) : %s",
                    len(manquants), ", ".join(sorted(manquants)))

    return records


# ── Insertion ─────────────────────────────────────────────────
def insert_records(conn, records, dry_run=False):
    if dry_run:
        log.info("[DRY-RUN] %d enregistrements non inseres", len(records))
        # Apercu
        from collections import Counter
        pays = Counter(r[1] for r in records)
        log.info("Exemple pays : %s",
                 ", ".join(f"{k}({v})" for k, v in list(pays.items())[:8]))
        return len(records)

    sql = """
        INSERT INTO ma.indicator_values
            (indicator_code, country_iso3, year, layer_id,
             raw_value, processed_value, method_version_id, quality_flag)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """

    # Compter avant
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM ma.indicator_values "
            "WHERE layer_id = %s AND indicator_code = %s",
            (LAYER_RAW, OSA_CODE)
        )
        before = cur.fetchone()[0]

    with conn.cursor() as cur:
        execute_batch(cur, sql, records, page_size=BATCH_SIZE)
    conn.commit()

    # Compter apres
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM ma.indicator_values "
            "WHERE layer_id = %s AND indicator_code = %s",
            (LAYER_RAW, OSA_CODE)
        )
        after = cur.fetchone()[0]

    inserted  = after - before
    conflicts = len(records) - inserted
    log.info("Inseres : %d | Conflits ignores : %d", inserted, conflicts)
    return inserted


# ── Bilan final ───────────────────────────────────────────────
def print_summary(conn):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT COUNT(*)                                     AS total,
                   COUNT(raw_value)                             AS non_null,
                   ROUND(COUNT(raw_value)*100.0/COUNT(*), 1)    AS coverage_pct,
                   ROUND(MIN(raw_value)::numeric, 3)            AS vmin,
                   ROUND(MAX(raw_value)::numeric, 3)            AS vmax,
                   ROUND(AVG(raw_value)::numeric, 3)            AS vmoy,
                   COUNT(DISTINCT country_iso3)                 AS n_pays
            FROM ma.indicator_values
            WHERE layer_id = %s AND indicator_code = %s
        """, (LAYER_RAW, OSA_CODE))
        row = cur.fetchone()
        if row:
            total, nn, cov, vmin, vmax, vmoy, npays = row
            log.info("Bilan MIL_DEP :")
            log.info("  Enregistrements : %d/%d (%.1f%%)", nn, total, cov)
            log.info("  Pays couverts   : %d", npays)
            log.info("  Plage           : [%.2f%%, %.2f%%] PIB", vmin or 0, vmax or 0)
            log.info("  Moyenne         : %.2f%% PIB", vmoy or 0)


# ── Point d'entree ────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="OSA -- Fetcher SIPRI Milex (MIL_DEP : depenses militaires % PIB)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Indicateur integre :
  MIL_DEP <- SIPRI "Share of GDP"  Depenses militaires / PIB

Valeurs brutes % PIB stockees en L1.
Normalisation [0,1] effectuee par normalize_indicator (pipeline L3).

Exemples :
  python fetcher_sipri_milex.py --file data/sipri/SIPRI-Milex-data-1949-2024.xlsx --dry-run
  python fetcher_sipri_milex.py --file data/sipri/SIPRI-Milex-data-1949-2024.xlsx
        """
    )
    parser.add_argument("--file",    required=True,
                        help="Chemin vers SIPRI-Milex-data-1949-2024.xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simulation sans ecriture en base")
    parser.add_argument("--output", choices=["csv", "db", "both"], default="both")
    args = parser.parse_args()

    log.info("=" * 55)
    log.info("OSA -- Fetcher SIPRI Milex")
    log.info("Fichier  : %s", args.file)
    log.info("Feuille  : %s", SHEET_NAME)
    log.info("Annees   : %d -> %d", YEAR_FROM, YEAR_TO)
    log.info("Dry-run  : %s", args.dry_run)
    log.info("=" * 55)

    rows = load_xlsx(args.file)

    conn = get_conn()
    try:
        african_iso3   = get_african_iso3(conn)
        method_version = get_method_version(conn)

        records = prepare_records(rows, african_iso3, method_version)

        if not records:
            log.warning("Aucun enregistrement a inserer")
            return

        n = insert_records(conn, records, args.dry_run)

        if not args.dry_run:
            print_summary(conn)

        log.info("=" * 55)
        log.info("SIPRI Milex termine | +%d valeurs inserees", n)
        log.info("=" * 55)

    finally:
        conn.close()


if __name__ == "__main__":
    main()
