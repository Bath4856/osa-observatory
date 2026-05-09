from openpyxl import load_workbook

wb = load_workbook(r'G:\osa-observatory\data\raw\gfw\global.xlsx', read_only=True)
ws = wb['Country tree cover loss']
all_rows = list(ws.iter_rows(values_only=True))
header = list(all_rows[0])

# Find tc_loss_ha_2020 column index
col_2020 = next(i for i,h in enumerate(header) if h and 'tc_loss_ha_2020' in str(h))

targets = ['Congo, Dem. Rep. of the','Democratic Republic of Congo',
           'Congo, Rep. of the','Republic of Congo','Congo',
           "Cote d'Ivoire",'Ivory Coast','Cote dIvoire',
           'Swaziland','Eswatini','Burkina Faso']

print('Pays / Seuil / tc_loss_2020:')
for row in all_rows[1:]:
    if row[0] in targets:
        print(f'  {row[0]} | seuil={row[1]} | 2020={row[col_2020]}')