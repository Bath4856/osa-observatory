from openpyxl import load_workbook

wb = load_workbook(r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2021-Africa_Summary-ERT.xlsx', read_only=True)
ws = wb['T3-Africa']
rows = list(ws.iter_rows(max_row=200, values_only=True))

# Afficher les lignes d'en-tete completes (rows 5-9)
print('=== EN-TETES ===')
for i in range(5, 10):
    print(f'Row {i}:', [v for v in rows[i] if v is not None])

# Afficher les 10 premiers pays avec toutes leurs valeurs
print('\n=== DONNEES PAYS (10 premiers) ===')
for i in range(10, 20):
    row = rows[i]
    vals = [(j, v) for j, v in enumerate(row) if v is not None and v != '--']
    if vals:
        print(f'{row[0]}: {vals}')

# Compter les colonnes non nulles
print('\n=== NOMBRE DE COLONNES ===')
print(f'Total colonnes: {len(rows[9])}')
print(f'Colonnes avec valeurs header: {sum(1 for v in rows[9] if v)}')

# Identifier les minerais couverts
print('\n=== MINERAIS IDENTIFIES (row 6-9) ===')
for col in range(0, len(rows[6]), 2):
    vals = [rows[r][col] for r in range(6,10) if rows[r][col]]
    if vals:
        print(f'Col {col}: {vals}')