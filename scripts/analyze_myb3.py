from openpyxl import load_workbook

wb = load_workbook(r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2021-Africa_Summary-ERT.xlsx', read_only=True)

# Analyser T1 et T2 aussi
for sheet in ['T1', 'T2', 'T3-Africa']:
    ws = wb[sheet]
    rows = list(ws.iter_rows(max_row=15, values_only=True))
    print(f'\n=== {sheet} ===')
    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        if vals:
            print(f'Row {i}: {vals[:8]}')

# Compter les lignes dans T3
ws3 = wb['T3-Africa']
all_rows = list(ws3.iter_rows(values_only=True))
print(f'\nTotal rows T3: {len(all_rows)}')

# Trouver les colonnes supplementaires (minerais T2)
print('\n=== TOUTES COLONNES T3 (headers) ===')
for col in range(0, 40, 2):
    if col < len(all_rows[9]):
        vals = [all_rows[r][col] for r in range(5,10) if col < len(all_rows[r]) and all_rows[r][col]]
        if vals:
            print(f'Col {col}: {vals}')