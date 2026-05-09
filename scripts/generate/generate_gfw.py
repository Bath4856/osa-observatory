from openpyxl import load_workbook

osa_names = {
    'Angola':'AGO','Benin':'BEN','Botswana':'BWA','Burkina Faso':'BFA',
    'Burundi':'BDI','Cameroon':'CMR','Central African Republic':'CAF',
    'Chad':'TCD','Comoros':'COM','Congo':'COG','Djibouti':'DJI',
    'Egypt':'EGY','Equatorial Guinea':'GNQ','Eritrea':'ERI','Eswatini':'SWZ',
    'Ethiopia':'ETH','Gabon':'GAB','Gambia':'GMB','Ghana':'GHA',
    'Guinea':'GIN','Guinea-Bissau':'GNB','Kenya':'KEN','Lesotho':'LSO',
    'Liberia':'LBR','Libya':'LBY','Madagascar':'MDG','Malawi':'MWI',
    'Mali':'MLI','Mauritania':'MRT','Mauritius':'MUS','Mozambique':'MOZ',
    'Namibia':'NAM','Niger':'NER','Nigeria':'NGA','Rwanda':'RWA',
    'Senegal':'SEN','Seychelles':'SYC','Sao Tome and Principe':'STP',
    'Sierra Leone':'SLE','Somalia':'SOM','South Africa':'ZAF',
    'South Sudan':'SSD','Sudan':'SDN','Togo':'TGO','Tunisia':'TUN',
    'Tanzania':'TZA','Uganda':'UGA','Zambia':'ZMB','Zimbabwe':'ZWE',
}

wb = load_workbook(r'G:\osa-observatory\data\global.xlsx', read_only=True)
ws = wb['Country tree cover loss']
all_rows = list(ws.iter_rows(values_only=True))
header = list(all_rows[0])

loss_cols = {h: i for i,h in enumerate(header)
             if h and 'tc_loss_ha_20' in str(h)
             and 2010 <= int(str(h).split('_')[-1]) <= 2024}

results = []
for row in all_rows[1:]:
    if row[1] != 30:
        continue
    country = row[0]
    if country not in osa_names:
        continue
    iso3 = osa_names[country]
    for col, idx in loss_cols.items():
        year = int(col.split('_')[-1])
        val = row[idx]
        if val and float(val) > 0:
            results.append((iso3, year, float(val)))

out = open(r'G:\osa-observatory\db\ingest_gfw_forest.sql', 'w', encoding='utf-8')
out.write("DO $$\nDECLARE v_ep INTEGER;\nBEGIN\n")
out.write("    SELECT id INTO v_ep FROM collect.provider_endpoints\n")
out.write("    WHERE endpoint_code = 'GFW_GLOBAL_XLS';\n\n")
out.write("    INSERT INTO collect.raw_data (endpoint_id, indicator_code, country_iso3, year, value_raw)\n")
out.write("    VALUES\n")

vals = [f"    (v_ep, 'ENV_DEF', '{iso3}', {year}, {val})"
        for iso3, year, val in results]
out.write(',\n'.join(vals) + ';\n')
out.write("END $$;\n")
out.close()
print(f"OK — {len(results)} lignes generees")