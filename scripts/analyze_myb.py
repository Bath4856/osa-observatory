from openpyxl import load_workbook
import pandas as pd

# Analyser tous les fichiers MYB disponibles
files = {
    2021: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2021-Africa_Summary-ERT.xlsx',
    2019: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2019-africa.xlsx',
    2016: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2016-africa-sum.xlsx',
}

for year, path in files.items():
    print(f'\n=== MYB {year} ===')
    wb = load_workbook(path, read_only=True)
    print('Sheets:', wb.sheetnames)
    # Find Africa sheet
    africa_sheet = next((s for s in wb.sheetnames
                        if 'africa' in s.lower() or 'T3' in s), None)
    if africa_sheet:
        ws = wb[africa_sheet]
        rows = list(ws.iter_rows(max_row=100, values_only=True))
        # Find header row with country names
        for i, row in enumerate(rows):
            if row[0] and 'Angola' in str(row[0]):
                print(f'  Premiere ligne pays: row {i}')
                # Get column headers
                header_row = rows[i-1]
                print(f'  Colonnes (sample): {[v for v in rows[9] if v]}')
                # Count countries
                countries = [r[0] for r in rows[i:] if r[0] and isinstance(r[0], str)
                            and r[0] not in ['Total','Source','Note']]
                print(f'  Nombre pays: {len(countries)}')
                print(f'  Pays: {countries[:10]}...')
                break