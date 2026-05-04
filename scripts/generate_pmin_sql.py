from openpyxl import load_workbook

osa_map = {
    'Angola':'AGO','Benin':'BEN','Botswana':'BWA','Burkina Faso':'BFA',
    'Burundi':'BDI','Cameroon':'CMR','Cape Verde':'CPV',
    'Central African Republic':'CAF','Chad':'TCD','Comoros':'COM',
    'Congo (Brazzaville)':'COG','Congo (Kinshasa)':'COD',
    'Djibouti':'DJI','Egypt':'EGY','Equatorial Guinea':'GNQ',
    'Eritrea':'ERI','Eswatini':'SWZ','Ethiopia':'ETH',
    'Gabon':'GAB','Gambia':'GMB','Ghana':'GHA',
    'Guinea':'GIN','Guinea-Bissau':'GNB','Kenya':'KEN',
    'Lesotho':'LSO','Liberia':'LBR','Libya':'LBY',
    'Madagascar':'MDG','Malawi':'MWI','Mali':'MLI',
    'Mauritania':'MRT','Mauritius':'MUS','Morocco':'MAR',
    'Mozambique':'MOZ','Namibia':'NAM','Niger':'NER',
    'Nigeria':'NGA','Rwanda':'RWA','Senegal':'SEN',
    'Seychelles':'SYC','Sierra Leone':'SLE','Somalia':'SOM',
    'South Africa':'ZAF','South Sudan':'SSD','Sudan':'SDN',
    'Tanzania':'TZA','Togo':'TGO','Tunisia':'TUN',
    'Uganda':'UGA','Zambia':'ZMB','Zimbabwe':'ZWE',
}

# Annees disponibles dans les fichiers MYB
files = {
    2021: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2021-Africa_Summary-ERT.xlsx',
    2019: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2019-africa.xlsx',
    2016: r'G:\osa-observatory\data\raw\pmin\usgs\myb3-2016-africa-sum.xlsx',
}

sheet_map = {
    2021: 'T3-Africa',
    2019: 'T3',
    2016: 'Table 3',
}

# col_idx -> (indicator_code, mineral)
col_map = {
    2:  'MIN_PRD_BAU',
    4:  'MIN_PRD_ALU',
    6:  'MIN_PRD_CHR',
    8:  'MIN_PRD_COB',
    10: 'MIN_PRD_COP',
    12: 'MIN_PRD_GOL',
    14: 'MIN_PRD_IRN',
    16: 'MIN_PRD_STL',
    18: 'MIN_PRD_MAN',
}

results = []

for year, path in files.items():
    try:
        wb = load_workbook(path, read_only=True)
        sheet = sheet_map[year]
        if sheet not in wb.sheetnames:
            sheet = next((s for s in wb.sheetnames
                         if 'africa' in s.lower() or 'T3' in s or 'Table 3' in s), None)
        if not sheet:
            print(f'Sheet non trouve pour {year}')
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(max_row=200, values_only=True))
        # Find first country row
        start = next(i for i,r in enumerate(rows)
                    if r[0] and r[0] in osa_map)
        for row in rows[start:]:
            country = row[0]
            if not country or country not in osa_map:
                continue
            iso3 = osa_map[country]
            for col, ind in col_map.items():
                if col < len(row):
                    val = row[col]
                    if val and val != '--' and str(val).strip() not in ('', '--', 'W'):
                        try:
                            fval = float(str(val).replace(',',''))
                            if fval > 0:
                                results.append((ind, iso3, year, fval))
                        except:
                            pass
        print(f'MYB {year}: {sum(1 for r in results if r[2]==year)} lignes extraites')
    except Exception as e:
        print(f'Erreur {year}: {e}')

# Generer le SQL
q = chr(39)
ep = '(SELECT id FROM collect.provider_endpoints WHERE endpoint_code = ' + q + 'USGS_MYB_AFRICA' + q + ')'

with open(r'G:\osa-observatory\db\ingest_usgs_myb.sql', 'w', encoding='utf-8') as f:
    f.write('-- PMIN F3 : Ingestion USGS MYB production mineraux\n')
    f.write('-- Annees : 2016, 2019, 2021\n')
    f.write('-- 9 mineraux critiques\n\n')
    for ind, iso3, year, val in results:
        f.write(f'INSERT INTO collect.raw_data (endpoint_id, indicator_code, country_iso3, year, value_raw) '
                f'VALUES ({ep}, {q}{ind}{q}, {q}{iso3}{q}, {year}, {val});\n')

print(f'Total: {len(results)} lignes -> ingest_usgs_myb.sql')